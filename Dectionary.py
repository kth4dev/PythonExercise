import openpyxl

ask = 'y'


def findwords(dbname):
    excel_file = openpyxl.load_workbook(dbname)
    sheet = excel_file.worksheets[0]
    result = ''
    for i in range(1, sheet.max_row):
        name = sheet.cell(i, 1).value
        if u_input == name:
            result = sheet.cell(i, 2).value
    if result != '':
        print(result)
    else:
        print('Not Found!')


while ask == 'y':
    type_input = input('Type(adj,noun,verb) = ')
    u_input = input('Name = ')
    if type_input == 'adj':
        findwords('ADJExcel.xlsx')
    elif type_input == 'noun':
        findwords('nounExcel.xlsx')
    elif type_input == 'verb':
        findwords('verbExcel.xlsx')
    ask = input("Find Words again? ( y or n)")
